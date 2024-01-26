import tkinter as tk
from tkinter import messagebox, simpledialog
from datetime import datetime
import openpyxl
from PIL import Image, ImageTk

class TodoApp:
    def __init__(self, master):
        self.master = master
        self.master.title("Todo List App")

        self.todo_list = []
        self.selected_todo_index = None
        self.load_todo_list_from_excel()
        self.create_widgets()

    def create_widgets(self):
        # Configuração da fonte e cor
        font_style = ("Helvetica", 12)
        bg_color = "#f2f2f2"
        button_bg_color = "#4caf50"
        button_fg_color = "white"

        # Configuração do ícone
        icon = Image.open("todo_icon.png")
        icon = icon.resize((50, 50), Image.ANTIALIAS if hasattr(Image, 'ANTIALIAS') else Image.BICUBIC)
        self.icon_image = ImageTk.PhotoImage(icon)
        self.master.iconphoto(True, self.icon_image)

        # Configuração da entrada de texto
        self.todo_entry = tk.Entry(self.master, font=font_style, width=30)
        self.todo_entry.pack(pady=10, padx=20)
        self.todo_entry.bind("<Return>", self.add_todo)

        # Configuração do botão "Add Todo"
        self.add_button = tk.Button(
            self.master, text="Add Todo", command=self.add_todo,
            font=font_style, bg=button_bg_color, fg=button_fg_color
        )
        self.add_button.pack(pady=5, padx=20)

        # Frame para a lista de tarefas
        list_frame = tk.Frame(self.master)
        list_frame.pack(pady=10, padx=20, side=tk.LEFT)

        # Configuração da lista de tarefas com barra de rolagem
        self.todo_listbox = tk.Listbox(
            list_frame, selectmode=tk.SINGLE, font=font_style,
            width=30, height=10, bg=bg_color
        )
        self.todo_listbox.pack(side=tk.LEFT)

        # Barra de rolagem vertical para a lista de tarefas
        scrollbar = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        scrollbar.config(command=self.todo_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.todo_listbox.config(yscrollcommand=scrollbar.set)

        # Configuração do botão "Edit Todo"
        self.edit_button = tk.Button(
            self.master, text="Edit Todo", command=self.edit_todo,
            font=font_style, bg=button_bg_color, fg=button_fg_color
        )
        self.edit_button.pack(pady=5, padx=20)

        # Configuração do botão "Delete Todo"
        self.delete_button = tk.Button(
            self.master, text="Delete Todo", command=self.delete_todo,
            font=font_style, bg="#ff3333", fg=button_fg_color
        )
        self.delete_button.pack(pady=5, padx=20)

        # Atualiza a lista de tarefas na interface gráfica
        self.update_todo_listbox()

    def add_todo(self, event=None):
        todo_text = self.todo_entry.get().strip()
        if todo_text:
            todo_id = len(self.todo_list) + 1
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            todo = {"id": todo_id, "text": todo_text, "timestamp": timestamp}
            self.todo_list.append(todo)
            self.update_todo_listbox()
            self.update_excel()
            self.todo_entry.delete(0, tk.END)
        else:
            messagebox.showwarning("Warning", "Please enter a todo.")

    def select_todo(self, event=None):
        selected_index = self.todo_listbox.curselection()
        if selected_index:
            self.selected_todo_index = selected_index[0]

    def edit_todo(self):
        if self.selected_todo_index is not None:
            edited_text = simpledialog.askstring("Edit Todo", "Enter new text:", initialvalue=self.todo_list[self.selected_todo_index]["text"])
            if edited_text is not None:
                self.todo_list[self.selected_todo_index]["text"] = edited_text
                self.update_todo_listbox()
                self.update_excel()

    def delete_todo(self):
        if self.selected_todo_index is not None:
            del self.todo_list[self.selected_todo_index]
            self.selected_todo_index = None
            self.update_todo_listbox()
            self.update_excel()

    def update_todo_listbox(self):
        self.todo_listbox.delete(0, tk.END)
        for todo in self.todo_list:
            self.todo_listbox.insert(tk.END, f"{todo['id']}: {todo['text']} ({todo['timestamp']})")

    def update_excel(self):
        workbook = openpyxl.load_workbook("todo_list.xlsx")
        sheet = workbook.active

        # Clear existing data
        for row in sheet.iter_rows(min_row=2, max_col=3, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        # Write updated data
        for row_num, todo in enumerate(self.todo_list, 2):
            sheet.cell(row=row_num, column=1, value=todo["id"])
            sheet.cell(row=row_num, column=2, value=todo["text"])
            sheet.cell(row=row_num, column=3, value=todo["timestamp"])

        workbook.save("todo_list.xlsx")

    def load_todo_list_from_excel(self):
        try:
            workbook = openpyxl.load_workbook("todo_list.xlsx")
            sheet = workbook.active
            for row in sheet.iter_rows(min_row=2, values_only=True):
                todo = {"id": row[0], "text": row[1], "timestamp": row[2]}
                self.todo_list.append(todo)
        except FileNotFoundError:
            pass

if __name__ == "__main__":
    root = tk.Tk()
    app = TodoApp(root)
    root.mainloop()
